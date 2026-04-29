import csv
import fnmatch
import html
import json
import os
import re
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


BASE_DIR = Path(__file__).resolve().parent

TRAINS_FILE = BASE_DIR / "trains.json"
LIVE_TRAINS_FILE = BASE_DIR / "live_trains.json"

LOCOS_FILE = BASE_DIR / "locos.json"
LOCO_HISTORY_FILE = BASE_DIR / "loco_history.json"
LOCO_EXPORT_FILE = BASE_DIR / "loco_export.csv"
LOCO_SUMMARY_FILE = BASE_DIR / "loco_summary.txt"

DOWNLOADS_DIR = BASE_DIR / "static" / "downloads"
LOCO_DATABASE_HTML = DOWNLOADS_DIR / "loco_database.html"
RECENTLY_ADDED_HTML = DOWNLOADS_DIR / "recently_added.html"
LOCO_NUMBERS_ONLY_HTML = DOWNLOADS_DIR / "loco_numbers_only.html"
LOCO_DATABASE_XLSX = DOWNLOADS_DIR / "loco_database.xlsx"
LOCO_NUMBERS_ONLY_XLSX = DOWNLOADS_DIR / "loco_numbers_only.xlsx"

BLOCKLIST_FILE = BASE_DIR / "blocklist.json"


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def iso_now() -> str:
    return utc_now().isoformat().replace("+00:00", "Z")


def ensure_dirs() -> None:
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)


def load_json(path: Path, default):
    if not path.exists():
        return deepcopy(default)
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return deepcopy(default)


def save_json(path: Path, payload) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def load_trains_payload() -> dict[str, Any]:
    if TRAINS_FILE.exists():
        payload = load_json(TRAINS_FILE, {})
    elif LIVE_TRAINS_FILE.exists():
        payload = load_json(LIVE_TRAINS_FILE, {})
    else:
        payload = {}

    if isinstance(payload, list):
        return {
            "lastUpdated": iso_now(),
            "trains": payload,
        }

    if isinstance(payload, dict):
        if isinstance(payload.get("trains"), list):
            return payload

        for key in ["items", "data", "features"]:
            if isinstance(payload.get(key), list):
                return {
                    "lastUpdated": payload.get("lastUpdated") or payload.get("updated") or iso_now(),
                    "trains": payload[key],
                }

    return {
        "lastUpdated": iso_now(),
        "trains": [],
    }


def load_existing_locos() -> list[dict[str, Any]]:
    payload = load_json(LOCOS_FILE, [])

    if isinstance(payload, list):
        return payload

    if isinstance(payload, dict):
        for key in ["locos", "items", "data"]:
            if isinstance(payload.get(key), list):
                return payload[key]

    return []


def load_blocklist() -> dict[str, list[str]]:
    payload = load_json(
        BLOCKLIST_FILE,
        {
            "blocked_locos": [],
            "blocked_routes": [],
            "blocked_descriptions": [],
            "blocked_operators": [],
        },
    )

    return {
        "blocked_locos": [str(x).strip() for x in payload.get("blocked_locos", []) if str(x).strip()],
        "blocked_routes": [str(x).strip() for x in payload.get("blocked_routes", []) if str(x).strip()],
        "blocked_descriptions": [str(x).strip() for x in payload.get("blocked_descriptions", []) if str(x).strip()],
        "blocked_operators": [str(x).strip() for x in payload.get("blocked_operators", []) if str(x).strip()],
    }


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_key(value: Any) -> str:
    text = norm_text(value).upper()
    text = re.sub(r"\s+", "", text)
    return text


def get_first(d: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        value = d.get(key)
        if value not in [None, ""]:
            return norm_text(value)
    return ""


def maybe_properties(train: Any) -> dict[str, Any]:
    if not isinstance(train, dict):
        return {}

    if isinstance(train.get("properties"), dict):
        merged = dict(train["properties"])
        for key, value in train.items():
            if key != "properties" and key not in merged:
                merged[key] = value
        return merged

    return train


def extract_loco_number(train: dict[str, Any]) -> str:
    candidates = [
        "loco_number",
        "locoNumber",
        "loco",
        "locomotive",
        "locomotive_number",
        "unit",
        "vehicle",
        "vehicle_id",
        "trKey",
        "name",
        "train_name",
        "trainName",
        "label",
        "id",
        "ID",
    ]

    raw = get_first(train, candidates)

    if not raw:
        return ""

    text = raw.strip()

    if "•" in text:
        text = text.split("•", 1)[0].strip()

    if "|" in text:
        text = text.split("|", 1)[0].strip()

    text = re.sub(r"^(LOCO|Loco|loco)\s*[:#-]?\s*", "", text).strip()

    match = re.search(r"[A-Z]{1,8}[- ]?\d{1,5}[A-Z]?", text.upper())
    if match:
        return match.group(0).replace(" ", "")

    match = re.search(r"\b\d{3,6}\b", text)
    if match:
        return match.group(0)

    return text.upper().replace(" ", "")


def extract_train_id(train: dict[str, Any]) -> str:
    return get_first(
        train,
        [
            "train_id",
            "trainId",
            "train_number",
            "trainNumber",
            "service",
            "service_number",
            "serviceNumber",
            "run",
            "route",
            "tt",
            "id",
        ],
    )


def extract_operator(train: dict[str, Any]) -> str:
    return get_first(
        train,
        [
            "current_operator",
            "operator",
            "operator_name",
            "rail_operator",
            "company",
            "owner",
        ],
    )


def extract_description(train: dict[str, Any]) -> str:
    return get_first(
        train,
        [
            "vehicle_description",
            "vehicleDescription",
            "description",
            "desc",
            "type",
            "class",
            "train_description",
            "service_description",
        ],
    )


def extract_route_text(train: dict[str, Any]) -> str:
    parts = [
        extract_train_id(train),
        get_first(train, ["origin", "from"]),
        get_first(train, ["destination", "to"]),
        get_first(train, ["route", "line", "path"]),
    ]
    return " ".join([p for p in parts if p])


def is_blocked_value(value: str, patterns: list[str], wildcard: bool = False) -> bool:
    if not value:
        return False

    value_l = value.lower()

    for pattern in patterns:
        p = str(pattern).strip()
        if not p:
            continue

        if wildcard or "*" in p or "?" in p:
            if fnmatch.fnmatch(value.upper(), p.upper()):
                return True
        else:
            if p.lower() in value_l:
                return True

    return False


def is_loco_blocked(loco_number: str, train: dict[str, Any], blocklist: dict[str, list[str]]) -> tuple[bool, str]:
    if is_blocked_value(loco_number, blocklist["blocked_locos"], wildcard=True):
        return True, "blocked_loco"

    route_text = extract_route_text(train)
    if is_blocked_value(route_text, blocklist["blocked_routes"]):
        return True, "blocked_route"

    description = extract_description(train)
    if is_blocked_value(description, blocklist["blocked_descriptions"]):
        return True, "blocked_description"

    operator = extract_operator(train)
    if is_blocked_value(operator, blocklist["blocked_operators"]):
        return True, "blocked_operator"

    return False, ""


def parse_date_sort(value: Any) -> datetime:
    text = norm_text(value)
    if not text:
        return datetime(1970, 1, 1, tzinfo=timezone.utc)

    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except Exception:
        pass

    for fmt in [
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%d %b %Y %H:%M",
        "%d %b %Y at %I:%M %p",
    ]:
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except Exception:
            pass

    return datetime(1970, 1, 1, tzinfo=timezone.utc)


def preferred_existing_date(existing: dict[str, Any]) -> str:
    for key in [
        "date_time_added",
        "Date/Time Added",
        "dateTimeAdded",
        "first_seen",
        "firstSeen",
        "added",
        "created_at",
    ]:
        value = existing.get(key)
        if value:
            return norm_text(value)
    return ""


def make_loco_record_from_train(train: dict[str, Any], now_iso: str) -> dict[str, Any] | None:
    loco_number = extract_loco_number(train)

    if not loco_number:
        return None

    operator = extract_operator(train)
    description = extract_description(train)
    train_id = extract_train_id(train)

    lat = get_first(train, ["lat", "latitude", "y"])
    lon = get_first(train, ["lon", "lng", "longitude", "x"])

    return {
        "loco_number": loco_number,
        "current_operator": operator,
        "vehicle_description": description,
        "train_id": train_id,
        "route": extract_route_text(train),
        "last_seen": now_iso,
        "date_time_added": now_iso,
        "lat": lat,
        "lon": lon,
        "source": "trains.json",
    }


def merge_locos(existing_locos: list[dict[str, Any]], trains: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    now_iso = iso_now()
    blocklist = load_blocklist()

    master: dict[str, dict[str, Any]] = {}

    for item in existing_locos:
        if not isinstance(item, dict):
            continue

        loco_number = get_first(item, ["loco_number", "Loco Number", "number", "loco", "id"])
        loco_key = norm_key(loco_number)

        if not loco_key:
            continue

        item = dict(item)
        item["loco_number"] = loco_number or loco_key

        if not preferred_existing_date(item):
            item["date_time_added"] = now_iso

        master[loco_key] = item

    new_added = []
    seen_this_run = 0

    for raw_train in trains:
        train = maybe_properties(raw_train)
        if not train:
            continue

        loco_number = extract_loco_number(train)
        loco_key = norm_key(loco_number)

        if not loco_key:
            continue

        blocked, reason = is_loco_blocked(loco_number, train, blocklist)
        if blocked:
            continue

        seen_this_run += 1

        new_record = make_loco_record_from_train(train, now_iso)
        if not new_record:
            continue

        if loco_key in master:
            existing = master[loco_key]

            original_added = preferred_existing_date(existing) or now_iso

            for key, value in new_record.items():
                if key == "date_time_added":
                    continue
                if value not in [None, ""]:
                    existing[key] = value

            existing["date_time_added"] = original_added
            existing["last_seen"] = now_iso
            master[loco_key] = existing
        else:
            master[loco_key] = new_record
            new_added.append(new_record)

    merged = list(master.values())

    merged.sort(
        key=lambda x: (
            parse_date_sort(x.get("date_time_added")).timestamp(),
            norm_key(x.get("loco_number")),
        ),
        reverse=True,
    )

    return merged, new_added, seen_this_run


def visible_locos(locos: list[dict[str, Any]]) -> list[dict[str, Any]]:
    blocklist = load_blocklist()
    output = []

    for loco in locos:
        loco_number = get_first(loco, ["loco_number", "Loco Number", "number", "loco", "id"])
        blocked, _ = is_loco_blocked(loco_number, loco, blocklist)
        if not blocked:
            output.append(loco)

    return output


def esc(value: Any) -> str:
    return html.escape(norm_text(value))


def display_date(value: Any) -> str:
    dt = parse_date_sort(value)
    if dt.year == 1970:
        return ""
    return dt.strftime("%d %b %Y %H:%M")


def loco_value(loco: dict[str, Any], keys: list[str]) -> str:
    return get_first(loco, keys)


def html_header(title: str, active: str, count: int, generated: str) -> str:
    def button(label: str, href: str, is_active: bool = False) -> str:
        cls = "btn active" if is_active else "btn"
        return f'<a class="{cls}" href="{href}">{html.escape(label)}</a>'

    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(title)}</title>
<style>
:root {{
  color-scheme: dark;
  --bg:#071222;
  --panel:#102039;
  --panel2:#0d1a2f;
  --line:#274569;
  --text:#eaf2ff;
  --muted:#9fb3d3;
  --accent:#58b6ff;
  --pill:#123f68;
}}
* {{ box-sizing:border-box; }}
html, body {{
  margin:0;
  padding:0;
  background:var(--bg);
  color:var(--text);
  font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Arial,sans-serif;
}}
body {{ padding:24px; }}
.card {{
  background:linear-gradient(180deg,var(--panel),var(--panel2));
  border:1px solid var(--line);
  border-radius:18px;
  padding:22px;
  margin-bottom:22px;
  box-shadow:0 12px 30px rgba(0,0,0,.25);
}}
h1 {{
  font-size:34px;
  line-height:1.1;
  margin:0 0 12px;
}}
.subtitle {{
  color:var(--muted);
  font-size:22px;
  line-height:1.35;
  margin-bottom:18px;
}}
.pill {{
  display:inline-block;
  background:var(--pill);
  border:1px solid #28557f;
  padding:10px 16px;
  border-radius:999px;
  margin:4px 8px 10px 0;
  font-size:18px;
}}
.nav {{
  display:flex;
  flex-wrap:wrap;
  gap:12px;
  margin-top:8px;
}}
.btn {{
  display:inline-block;
  text-decoration:none;
  color:var(--text);
  border:1px solid var(--line);
  border-radius:14px;
  padding:14px 20px;
  font-size:20px;
  background:#081426;
}}
.btn.active {{
  background:var(--accent);
  color:#03101d;
  font-weight:800;
}}
.table-wrap {{
  overflow-x:auto;
  border:1px solid var(--line);
  border-radius:18px;
}}
table {{
  width:100%;
  min-width:900px;
  border-collapse:collapse;
  background:#0d1b31;
}}
th, td {{
  border-bottom:1px solid var(--line);
  padding:16px;
  text-align:left;
  vertical-align:top;
  font-size:18px;
}}
th {{
  font-size:20px;
  background:#10213b;
  position:sticky;
  top:0;
  z-index:2;
}}
td.muted {{ color:var(--muted); }}
.raw {{ color:var(--muted); font-size:15px; margin-top:6px; }}
.numbers {{
  columns:5 130px;
  column-gap:32px;
  font-size:20px;
  line-height:1.9;
}}
.number-item {{
  break-inside:avoid;
  border-bottom:1px solid rgba(255,255,255,.08);
}}
@media (max-width:700px) {{
  body {{ padding:14px; }}
  h1 {{ font-size:28px; }}
  .subtitle {{ font-size:20px; }}
  th, td {{ font-size:17px; padding:14px; }}
  .btn {{ font-size:18px; }}
}}
</style>
</head>
<body>
<div class="card">
  <h1>{esc(title)}</h1>
  <div class="subtitle">Newest locos first, based on the Date/Time Added field.</div>
  <div>
    <span class="pill">Visible locos: {count}</span>
    <span class="pill">Generated: {esc(generated)}</span>
  </div>
  <div class="nav">
    {button("Full database", "loco_database.html", active == "full")}
    {button("Recently added", "recently_added.html", active == "recent")}
    {button("Numbers only", "loco_numbers_only.html", active == "numbers")}
    {button("Download workbook", "loco_database.xlsx")}
    {button("Download numbers workbook", "loco_numbers_only.xlsx")}
  </div>
</div>
"""


def html_footer() -> str:
    return "</body></html>\n"


def generate_database_html(locos: list[dict[str, Any]], generated_label: str) -> None:
    rows = []

    for loco in locos:
        loco_number = loco_value(loco, ["loco_number", "Loco Number", "number", "loco"])
        operator = loco_value(loco, ["current_operator", "Current Operator", "operator"])
        description = loco_value(loco, ["vehicle_description", "Vehicle Description", "description"])
        train_id = loco_value(loco, ["train_id", "Train ID", "service", "route"])
        added = loco_value(loco, ["date_time_added", "Date/Time Added", "first_seen", "added"])
        last_seen = loco_value(loco, ["last_seen", "Last Seen"])

        rows.append(
            f"""
<tr>
  <td><strong>{esc(loco_number)}</strong></td>
  <td>{esc(operator)}</td>
  <td>{esc(description)}</td>
  <td>{esc(train_id)}</td>
  <td><strong>{esc(display_date(added))}</strong><div class="raw">Raw: {esc(added)}</div></td>
  <td>{esc(display_date(last_seen))}<div class="raw">Raw: {esc(last_seen)}</div></td>
</tr>
"""
        )

    html_text = html_header("RailOps Loco Database", "full", len(locos), generated_label)
    html_text += """
<div class="card table-wrap">
<table>
<thead>
<tr>
  <th>Loco Number</th>
  <th>Current Operator</th>
  <th>Vehicle Description</th>
  <th>Train/Service</th>
  <th>Date/Time Added</th>
  <th>Last Seen</th>
</tr>
</thead>
<tbody>
"""
    html_text += "\n".join(rows)
    html_text += """
</tbody>
</table>
</div>
"""
    html_text += html_footer()

    LOCO_DATABASE_HTML.write_text(html_text, encoding="utf-8")


def generate_recent_html(locos: list[dict[str, Any]], generated_label: str, limit: int = 300) -> None:
    recent = sorted(
        locos,
        key=lambda x: parse_date_sort(loco_value(x, ["date_time_added", "Date/Time Added", "first_seen", "added"])),
        reverse=True,
    )[:limit]

    rows = []

    for loco in recent:
        loco_number = loco_value(loco, ["loco_number", "Loco Number", "number", "loco"])
        operator = loco_value(loco, ["current_operator", "Current Operator", "operator"])
        description = loco_value(loco, ["vehicle_description", "Vehicle Description", "description"])
        added = loco_value(loco, ["date_time_added", "Date/Time Added", "first_seen", "added"])

        rows.append(
            f"""
<tr>
  <td><strong>{esc(loco_number)}</strong></td>
  <td>{esc(operator)}</td>
  <td>{esc(description)}</td>
  <td><strong>{esc(display_date(added))}</strong><div class="raw">Raw: {esc(added)}</div></td>
</tr>
"""
        )

    html_text = html_header("RailOps Recently Added Locos", "recent", len(locos), generated_label)
    html_text += """
<div class="card table-wrap">
<table>
<thead>
<tr>
  <th>Loco Number</th>
  <th>Current Operator</th>
  <th>Vehicle Description</th>
  <th>Date/Time Added</th>
</tr>
</thead>
<tbody>
"""
    html_text += "\n".join(rows)
    html_text += """
</tbody>
</table>
</div>
"""
    html_text += html_footer()

    RECENTLY_ADDED_HTML.write_text(html_text, encoding="utf-8")


def generate_numbers_html(locos: list[dict[str, Any]], generated_label: str) -> None:
    numbers = sorted(
        {
            loco_value(loco, ["loco_number", "Loco Number", "number", "loco"])
            for loco in locos
            if loco_value(loco, ["loco_number", "Loco Number", "number", "loco"])
        },
        key=lambda x: (re.sub(r"\d+", "", x), [int(n) for n in re.findall(r"\d+", x)] or [0], x),
    )

    items = "\n".join(f'<div class="number-item">{esc(number)}</div>' for number in numbers)

    html_text = html_header("RailOps Loco Numbers Only", "numbers", len(locos), generated_label)
    html_text += f"""
<div class="card">
  <div class="numbers">
    {items}
  </div>
</div>
"""
    html_text += html_footer()

    LOCO_NUMBERS_ONLY_HTML.write_text(html_text, encoding="utf-8")


def generate_csv(locos: list[dict[str, Any]]) -> None:
    fields = [
        "loco_number",
        "current_operator",
        "vehicle_description",
        "train_id",
        "route",
        "date_time_added",
        "last_seen",
        "lat",
        "lon",
        "source",
    ]

    with LOCO_EXPORT_FILE.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()

        for loco in locos:
            row = {field: loco.get(field, "") for field in fields}
            writer.writerow(row)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="10213B")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="111111")
    thin = Side(style="thin", color="C7D3E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = body_font

    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        max_len = 10
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 45))
        ws.column_dimensions[col_letter].width = max_len + 2


def generate_xlsx(locos: list[dict[str, Any]]) -> None:
    if Workbook is None:
        print("openpyxl not installed. Skipping workbook generation.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Loco Database"

    headers = [
        "Loco Number",
        "Current Operator",
        "Vehicle Description",
        "Train/Service",
        "Route",
        "Date/Time Added",
        "Last Seen",
        "Latitude",
        "Longitude",
        "Source",
    ]

    ws.append(headers)

    for loco in locos:
        ws.append(
            [
                loco_value(loco, ["loco_number", "Loco Number", "number", "loco"]),
                loco_value(loco, ["current_operator", "Current Operator", "operator"]),
                loco_value(loco, ["vehicle_description", "Vehicle Description", "description"]),
                loco_value(loco, ["train_id", "Train ID", "service"]),
                loco_value(loco, ["route"]),
                loco_value(loco, ["date_time_added", "Date/Time Added", "first_seen", "added"]),
                loco_value(loco, ["last_seen", "Last Seen"]),
                loco_value(loco, ["lat", "latitude"]),
                loco_value(loco, ["lon", "lng", "longitude"]),
                loco_value(loco, ["source"]),
            ]
        )

    style_sheet(ws)
    wb.save(LOCO_DATABASE_XLSX)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Numbers Only"
    ws2.append(["Loco Number"])

    numbers = sorted(
        {
            loco_value(loco, ["loco_number", "Loco Number", "number", "loco"])
            for loco in locos
            if loco_value(loco, ["loco_number", "Loco Number", "number", "loco"])
        }
    )

    for number in numbers:
        ws2.append([number])

    style_sheet(ws2)
    wb2.save(LOCO_NUMBERS_ONLY_XLSX)


def generate_summary(
    trains_count: int,
    existing_before: int,
    merged_count: int,
    new_added_count: int,
    seen_this_run: int,
    generated_iso: str,
) -> None:
    text = f"""RailOps Loco Database Summary
Generated UTC: {generated_iso}

Source trains this run: {trains_count}
Locos seen this run: {seen_this_run}
Existing locos before merge: {existing_before}
New locos added this run: {new_added_count}
Final visible/master locos: {merged_count}

Storage mode: GitHub committed database files
Rule: Existing locos are kept even if missing from a scrape.
Blocklist file: blocklist.json
"""
    LOCO_SUMMARY_FILE.write_text(text, encoding="utf-8")


def main() -> None:
    ensure_dirs()

    generated_iso = iso_now()
    generated_label = datetime.now().strftime("%d %b %Y %H:%M")

    trains_payload = load_trains_payload()
    trains = trains_payload.get("trains", [])
    if not isinstance(trains, list):
        trains = []

    existing = load_existing_locos()
    existing_before = len(existing)

    merged, new_added, seen_this_run = merge_locos(existing, trains)
    visible = visible_locos(merged)

    save_json(LOCOS_FILE, visible)

    history = load_json(LOCO_HISTORY_FILE, [])
    if not isinstance(history, list):
        history = []

    history.insert(
        0,
        {
            "generated": generated_iso,
            "source_trains": len(trains),
            "seen_this_run": seen_this_run,
            "existing_before": existing_before,
            "new_added": len(new_added),
            "final_count": len(visible),
            "new_loco_numbers": [
                loco_value(loco, ["loco_number", "Loco Number", "number", "loco"])
                for loco in new_added
            ],
        },
    )

    history = history[:500]
    save_json(LOCO_HISTORY_FILE, history)

    generate_database_html(visible, generated_label)
    generate_recent_html(visible, generated_label)
    generate_numbers_html(visible, generated_label)
    generate_csv(visible)
    generate_xlsx(visible)
    generate_summary(
        trains_count=len(trains),
        existing_before=existing_before,
        merged_count=len(visible),
        new_added_count=len(new_added),
        seen_this_run=seen_this_run,
        generated_iso=generated_iso,
    )

    print("RailOps loco database generated.")
    print(f"Source trains: {len(trains)}")
    print(f"Existing locos before merge: {existing_before}")
    print(f"Seen this run: {seen_this_run}")
    print(f"New locos added: {len(new_added)}")
    print(f"Final visible locos: {len(visible)}")
    print(f"Wrote: {LOCOS_FILE}")
    print(f"Wrote: {RECENTLY_ADDED_HTML}")
    print(f"Wrote: {LOCO_DATABASE_HTML}")
    print(f"Wrote: {LOCO_NUMBERS_ONLY_HTML}")


if __name__ == "__main__":
    main()