import os
import json
import re
import html
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break

LOCOS_FILE = "locos.json"
BLOCKED_FILE = "blocked_locos.txt"
BLOCKED_DESCRIPTIONS_FILE = "blocked_descriptions.txt"
DOWNLOAD_DIR = os.path.join("static", "downloads")
XLSX_FILE = os.path.join(DOWNLOAD_DIR, "loco_database.xlsx")
NUMBERS_ONLY_FILE = os.path.join(DOWNLOAD_DIR, "loco_numbers_only.xlsx")
HTML_DATABASE_FILE = os.path.join(DOWNLOAD_DIR, "loco_database.html")
HTML_RECENT_FILE = os.path.join(DOWNLOAD_DIR, "recently_added.html")
HTML_NUMBERS_FILE = os.path.join(DOWNLOAD_DIR, "loco_numbers_only.html")

CM_TO_INCH = 1 / 2.54
MARGIN_1_5CM = 1.5 * CM_TO_INCH
NUMBERS_ONLY_COLUMNS = 8
NUMBERS_ONLY_COL_WIDTH = 12
NUMBERS_ONLY_ROWS_PER_PAGE = 56
NUMBERS_ONLY_PAGE_GAP = 1

SKIP_PREFIXES = (
    "ARROWMARKERSSOURCE_",
    "MARKERSOURCE_",
    "REGTRAINSSOURCE_",
    "UNREGTRAINSSOURCE_",
    "TRAINSOURCE_",
)


def is_real_loco_id(value):
    loco = normalize_loco(value)
    if not loco:
        return False
    return not any(prefix in loco for prefix in SKIP_PREFIXES)


def load_json(filename):
    if not os.path.exists(filename):
        return {}
    with open(filename, 'r', encoding='utf-8') as f:
        return json.load(f)


def normalize_loco(value):
    if value is None:
        return ""
    return str(value).strip().upper()


def safe_text(value):
    if value is None:
        return ""
    return str(value).strip()


def load_blocked_loco_rules():
    exact = set()
    prefixes = []
    if not os.path.exists(BLOCKED_FILE):
        return exact, prefixes
    with open(BLOCKED_FILE, 'r', encoding='utf-8') as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line or line.startswith('#'):
                continue
            value = normalize_loco(line)
            if value.endswith('*'):
                prefix = value[:-1]
                if prefix:
                    prefixes.append(prefix)
            else:
                exact.add(value)
    prefixes.sort(key=len, reverse=True)
    return exact, prefixes


def loco_is_blocked(loco_id, blocked_exact, blocked_prefixes):
    loco = normalize_loco(loco_id)
    if not loco:
        return False
    if loco in blocked_exact:
        return True
    return any(loco.startswith(prefix) for prefix in blocked_prefixes)


def load_blocked_descriptions():
    blocked = []
    if not os.path.exists(BLOCKED_DESCRIPTIONS_FILE):
        return blocked
    with open(BLOCKED_DESCRIPTIONS_FILE, 'r', encoding='utf-8') as f:
        for raw_line in f:
            line = raw_line.strip().lower()
            if not line or line.startswith('#'):
                continue
            blocked.append(line)
    return blocked


def description_is_blocked(description, blocked_descriptions):
    desc = safe_text(description).lower()
    if not desc or not blocked_descriptions:
        return False
    return any(term in desc for term in blocked_descriptions)


def natural_sort_key(value):
    text = normalize_loco(value)
    return [int(part) if part.isdigit() else part for part in re.split(r'(\d+)', text)]


def parse_datetime_sort(value):
    text = safe_text(value)
    if not text:
        return datetime.min
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return datetime.min


def get_visible_locos(locos):
    blocked_exact, blocked_prefixes = load_blocked_loco_rules()
    blocked_descriptions = load_blocked_descriptions()
    visible = []
    for loco_number, data in locos.items():
        if loco_is_blocked(loco_number, blocked_exact, blocked_prefixes):
            continue
        if not isinstance(data, dict):
            continue
        if not is_real_loco_id(loco_number):
            continue
        description = data.get('vehicle_description') or data.get('last_description') or ''
        if description_is_blocked(description, blocked_descriptions):
            continue
        visible.append((normalize_loco(loco_number), data))
    return visible


def autosize_columns(ws, widths):
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def apply_print_settings(ws, orientation='portrait', repeat_header=True):
    ws.page_margins = PageMargins(
        left=MARGIN_1_5CM,
        right=MARGIN_1_5CM,
        top=MARGIN_1_5CM,
        bottom=MARGIN_1_5CM,
        header=0.3,
        footer=0.3,
    )
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    if repeat_header:
        ws.print_title_rows = '1:1'


def make_row(loco_number, data):
    return [
        loco_number,
        safe_text(data.get('current_operator')),
        safe_text(data.get('vehicle_description') or data.get('last_description')),
        safe_text(data.get('date_time_added') or data.get('first_seen')),
    ]


def write_header(ws, headers, bold, center):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center


def write_numbers_pages(ws, running_order, center, columns=NUMBERS_ONLY_COLUMNS, rows_per_page=NUMBERS_ONLY_ROWS_PER_PAGE):
    loco_numbers = [loco_number for loco_number, _data in running_order]

    for col in range(1, columns + 1):
        ws.column_dimensions[get_column_letter(col)].width = NUMBERS_ONLY_COL_WIDTH

    row_pointer = 1
    total_per_page = columns * rows_per_page

    for page_start in range(0, len(loco_numbers), total_per_page):
        page_items = loco_numbers[page_start:page_start + total_per_page]

        for index, loco in enumerate(page_items):
            column_offset = index // rows_per_page
            row_offset = index % rows_per_page
            cell = ws.cell(row=row_pointer + row_offset, column=1 + column_offset, value=loco)
            cell.alignment = center

        next_page_row = row_pointer + rows_per_page
        if page_start + total_per_page < len(loco_numbers):
            ws.row_breaks.append(Break(id=next_page_row - 1))
            row_pointer = next_page_row + NUMBERS_ONLY_PAGE_GAP

    apply_print_settings(ws, orientation='portrait', repeat_header=False)


def format_dt_display(value):
    text = safe_text(value)
    if not text:
        return ""
    dt = parse_datetime_sort(text)
    if dt == datetime.min:
        return text
    return dt.strftime("%d %b %Y %H:%M")


def build_numbers_columns_html(running_order, columns=NUMBERS_ONLY_COLUMNS):
    locos = [loco for loco, _data in running_order]
    if not locos:
        return '<div class="empty">No loco numbers available.</div>'

    rows = (len(locos) + columns - 1) // columns
    groups = []
    for col in range(columns):
        start = col * rows
        end = start + rows
        chunk = locos[start:end]
        if chunk:
            items = "".join(f"<li>{html.escape(item)}</li>" for item in chunk)
            groups.append(f"<ol class=\"number-col\" start=\"{start + 1}\">{items}</ol>")
    return f"<div class=\"numbers-columns\">{''.join(groups)}</div>"


def render_html_page(title, subtitle, body, stats, active_tab):
    updated = html.escape(stats.get('generated_at', ''))
    visible_count = stats.get('visible_count', 0)
    nav = [
        ('database', 'Full database', 'loco_database.html'),
        ('recent', 'Recently added', 'recently_added.html'),
        ('numbers', 'Numbers only', 'loco_numbers_only.html'),
        ('xlsx', 'Download workbook', 'loco_database.xlsx'),
        ('numbers_xlsx', 'Download numbers workbook', 'loco_numbers_only.xlsx'),
    ]
    nav_html = ''.join(
        f'<a class="nav-link {"active" if key == active_tab else ""}" href="{html.escape(link)}">{html.escape(label)}</a>'
        for key, label, link in nav
    )
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    :root {{
      --bg: #08111f;
      --panel: #101c30;
      --panel-2: #16243b;
      --text: #e9f1ff;
      --muted: #9fb1cf;
      --accent: #5bb5ff;
      --border: #28415f;
      --chip: #0f3154;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, Helvetica, sans-serif; background: var(--bg); color: var(--text); }}
    .wrap {{ max-width: 1280px; margin: 0 auto; padding: 16px; }}
    .hero {{ background: linear-gradient(180deg, var(--panel), var(--panel-2)); border: 1px solid var(--border); border-radius: 18px; padding: 18px; margin-bottom: 16px; }}
    h1 {{ margin: 0 0 6px; font-size: 1.7rem; }}
    .subtitle {{ color: var(--muted); margin-bottom: 10px; }}
    .stats {{ display: flex; gap: 10px; flex-wrap: wrap; margin-top: 8px; }}
    .chip {{ background: var(--chip); border: 1px solid var(--border); border-radius: 999px; padding: 8px 12px; color: var(--text); font-size: 0.92rem; }}
    .nav {{ display: flex; gap: 10px; flex-wrap: wrap; margin-top: 14px; }}
    .nav-link {{ text-decoration: none; color: var(--text); background: #0c1525; border: 1px solid var(--border); border-radius: 12px; padding: 10px 12px; }}
    .nav-link.active {{ background: var(--accent); color: #00111f; border-color: var(--accent); font-weight: 700; }}
    .panel {{ background: var(--panel); border: 1px solid var(--border); border-radius: 18px; padding: 14px; overflow: hidden; }}
    .table-wrap {{ overflow: auto; border-radius: 12px; border: 1px solid var(--border); }}
    table {{ width: 100%; border-collapse: collapse; min-width: 720px; }}
    th, td {{ padding: 10px 12px; border-bottom: 1px solid var(--border); text-align: left; vertical-align: top; }}
    th {{ position: sticky; top: 0; background: #112039; z-index: 1; }}
    tr:nth-child(even) td {{ background: rgba(255,255,255,0.02); }}
    .numbers-columns {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(100px, 1fr)); gap: 14px; }}
    .number-col {{ margin: 0; padding-left: 28px; background: #0c1525; border: 1px solid var(--border); border-radius: 12px; min-height: 100%; }}
    .number-col li {{ padding: 6px 8px 6px 0; border-bottom: 1px solid rgba(255,255,255,0.05); }}
    .number-col li:last-child {{ border-bottom: 0; }}
    .hint {{ color: var(--muted); font-size: 0.92rem; margin-top: 10px; }}
    .empty {{ padding: 16px; color: var(--muted); }}
    @media (max-width: 760px) {{
      .wrap {{ padding: 10px; }}
      h1 {{ font-size: 1.35rem; }}
      .hero, .panel {{ border-radius: 14px; }}
      table {{ min-width: 620px; }}
      .numbers-columns {{ grid-template-columns: repeat(2, minmax(120px, 1fr)); }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <h1>{html.escape(title)}</h1>
      <div class="subtitle">{html.escape(subtitle)}</div>
      <div class="stats">
        <span class="chip">Visible locos: {visible_count}</span>
        <span class="chip">Updated: {updated}</span>
      </div>
      <nav class="nav">{nav_html}</nav>
    </section>
    <section class="panel">{body}</section>
  </div>
</body>
</html>
"""


def build_database_table(rows):
    if not rows:
        return '<div class="empty">No locos available.</div>'
    body_rows = ''.join(
        '<tr>' + ''.join(f'<td>{html.escape(value)}</td>' for value in row) + '</tr>'
        for row in rows
    )
    return (
        '<div class="table-wrap"><table><thead><tr>'
        '<th>Loco Number</th><th>Current Operator</th><th>Vehicle Description</th><th>Date/Time Added</th>'
        '</tr></thead><tbody>' + body_rows + '</tbody></table></div>'
        '<div class="hint">This page is rebuilt automatically every time the loco updater runs.</div>'
    )


def write_text_file(path, content):
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content)


def build_html_exports(running_order, recently_added, stats):
    running_rows = [make_row(loco_number, data) for loco_number, data in running_order]
    recent_rows = [make_row(loco_number, data) for loco_number, data in recently_added]

    running_rows = [[row[0], row[1], row[2], format_dt_display(row[3])] for row in running_rows]
    recent_rows = [[row[0], row[1], row[2], format_dt_display(row[3])] for row in recent_rows]

    write_text_file(
        HTML_DATABASE_FILE,
        render_html_page(
            title='RailOps Loco Database',
            subtitle='Full loco list in running order. Handy for phone viewing and quick searching in-browser.',
            body=build_database_table(running_rows),
            stats=stats,
            active_tab='database',
        )
    )

    write_text_file(
        HTML_RECENT_FILE,
        render_html_page(
            title='RailOps Recently Added Locos',
            subtitle='Newest locos first, based on the Date/Time Added field.',
            body=build_database_table(recent_rows),
            stats=stats,
            active_tab='recent',
        )
    )

    numbers_body = (
        build_numbers_columns_html(running_order, columns=NUMBERS_ONLY_COLUMNS)
        + '<div class="hint">Numbers run down each column first, then continue into the next column.</div>'
    )
    write_text_file(
        HTML_NUMBERS_FILE,
        render_html_page(
            title='RailOps Loco Numbers Only',
            subtitle='Numbers-only mobile page in running order.',
            body=numbers_body,
            stats=stats,
            active_tab='numbers',
        )
    )


def build_workbooks(locos):
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    visible_locos = get_visible_locos(locos)
    blocked_exact, blocked_prefixes = load_blocked_loco_rules()
    blocked_descriptions = sorted(load_blocked_descriptions())

    blocked_exact_sorted = sorted(blocked_exact, key=natural_sort_key)
    blocked_prefix_sorted = sorted(blocked_prefixes, key=natural_sort_key)

    running_order = sorted(visible_locos, key=lambda x: natural_sort_key(x[0]))
    recently_added = sorted(
        visible_locos,
        key=lambda x: parse_datetime_sort(x[1].get('date_time_added') or x[1].get('first_seen')),
        reverse=True,
    )

    headers = ["Loco Number", "Current Operator", "Vehicle Description", "Date/Time Added"]
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    widths = {1: 18, 2: 24, 3: 40, 4: 22}

    wb = Workbook()
    wb.remove(wb.active)

    ws_locos = wb.create_sheet("Locos")
    ws_recent = wb.create_sheet("Recently Added")
    ws_numbers = wb.create_sheet("Numbers Only")
    ws_blocked = wb.create_sheet("Blocked")
    ws_blocked_desc = wb.create_sheet("Blocked Descriptions")
    ws_info = wb.create_sheet("Instructions")

    write_header(ws_locos, headers, bold, center)
    for loco_number, data in running_order:
        ws_locos.append(make_row(loco_number, data))
    autosize_columns(ws_locos, widths)
    ws_locos.freeze_panes = 'A2'
    apply_print_settings(ws_locos, orientation='portrait')

    write_header(ws_recent, headers, bold, center)
    for loco_number, data in recently_added:
        ws_recent.append(make_row(loco_number, data))
    autosize_columns(ws_recent, widths)
    ws_recent.freeze_panes = 'A2'
    apply_print_settings(ws_recent, orientation='portrait')

    write_numbers_pages(ws_numbers, running_order, center)

    ws_blocked.append(["Blocked Loco Rule"])
    ws_blocked['A1'].font = bold
    ws_blocked['A1'].alignment = center
    ws_blocked.column_dimensions['A'].width = 24
    for loco in blocked_exact_sorted:
        ws_blocked.append([loco])
    for prefix in blocked_prefix_sorted:
        ws_blocked.append([f"{prefix}*"])
    apply_print_settings(ws_blocked, orientation='portrait')

    ws_blocked_desc.append(['Blocked Vehicle Description'])
    ws_blocked_desc['A1'].font = bold
    ws_blocked_desc['A1'].alignment = center
    ws_blocked_desc.column_dimensions['A'].width = 36
    for desc in blocked_descriptions:
        ws_blocked_desc.append([desc])
    apply_print_settings(ws_blocked_desc, orientation='portrait')

    lines = [
        "This workbook is rebuilt automatically by the loco updater.",
        "",
        "Locos: full loco list in running order.",
        "Recently Added: newest locos first, based on Date/Time Added.",
        "Numbers Only: loco numbers only, printed down each column first, then across the same page before moving to the next page.",
        "Blocked: add one loco rule per line in blocked_locos.txt.",
        "Use an exact rule like NR74 or a prefix rule like TNSW*.",
        "Blocked Descriptions: partial-match vehicle description filters from blocked_descriptions.txt.",
        "",
        "Print margins on the print sheets are set to about 1.5 cm.",
        f"Numbers Only layout uses {NUMBERS_ONLY_COLUMNS} columns and about {NUMBERS_ONLY_ROWS_PER_PAGE} rows per printed page.",
        "Phone-friendly HTML pages are generated automatically too:",
        "- /downloads/loco_database.html",
        "- /downloads/recently_added.html",
        "- /downloads/loco_numbers_only.html",
        "Download files:",
        "- /downloads/loco_database.xlsx",
        "- /downloads/loco_numbers_only.xlsx",
    ]
    for line in lines:
        ws_info.append([line])
    ws_info.column_dimensions['A'].width = 110
    apply_print_settings(ws_info, orientation='portrait', repeat_header=False)

    wb.save(XLSX_FILE)

    wb_numbers = Workbook()
    ws_only = wb_numbers.active
    ws_only.title = "Loco Numbers"
    write_numbers_pages(ws_only, running_order, center)
    wb_numbers.save(NUMBERS_ONLY_FILE)

    stats = {
        'visible_count': len(running_order),
        'blocked_exact_count': len(blocked_exact_sorted),
        'blocked_prefix_count': len(blocked_prefix_sorted),
        'blocked_description_count': len(blocked_descriptions),
        'recent_count': len(recently_added),
        'generated_at': datetime.now().strftime('%d %b %Y %H:%M'),
    }
    build_html_exports(running_order, recently_added, stats)

    return stats



def main():
    locos = load_json(LOCOS_FILE)
    if not isinstance(locos, dict):
        raise SystemExit("locos.json is missing or invalid")
    stats = build_workbooks(locos)
    print(f"✅ Workbook saved to {XLSX_FILE}")
    print(f"✅ Numbers-only workbook saved to {NUMBERS_ONLY_FILE}")
    print(f"✅ HTML database saved to {HTML_DATABASE_FILE}")
    print(f"✅ HTML recent page saved to {HTML_RECENT_FILE}")
    print(f"✅ HTML numbers page saved to {HTML_NUMBERS_FILE}")
    print(f"🚂 Visible locos: {stats['visible_count']}")
    print(f"🆕 Recently added rows: {stats['recent_count']}")
    print(f"🚫 Blocked exact locos listed: {stats['blocked_exact_count']}")
    print(f"🚫 Blocked loco prefixes listed: {stats['blocked_prefix_count']}")
    print(f"🚫 Blocked descriptions listed: {stats['blocked_description_count']}")


if __name__ == '__main__':
    main()
