import os
import json
import sqlite3
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# =========================
# Paths
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "locomotives.db")
BLOCKED_PATH = os.path.join(BASE_DIR, "blocked_locos.txt")
DOWNLOAD_DIR = os.path.join(BASE_DIR, "static", "downloads")
XLSX_PATH = os.path.join(DOWNLOAD_DIR, "loco_database.xlsx")

os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# =========================
# Database setup
# =========================
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS locos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            loco_number TEXT NOT NULL UNIQUE,
            current_operator TEXT,
            vehicle_description TEXT,
            date_time_added TEXT NOT NULL
        )
    """)

    conn.commit()
    conn.close()

# =========================
# Blocked list
# =========================
def load_blocked_locos():
    blocked = set()

    if not os.path.exists(BLOCKED_PATH):
        return blocked

    with open(BLOCKED_PATH, "r", encoding="utf-8") as f:
        for line in f:
            loco = line.strip().upper()
            if loco:
                blocked.add(loco)

    return blocked

# =========================
# Normalise values
# =========================
def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()

def normalise_loco(value):
    return clean_text(value).upper()

# =========================
# Replace this with your real scraper
# =========================
def scrape_latest_locos():
    """
    Replace this function with your actual scraper logic.
    It must return a list of dicts with:
      - loco_number
      - current_operator
      - vehicle_description
    """

    # Example/demo records only
    return [
        {
            "loco_number": "GWB106",
            "current_operator": "Aurizon",
            "vehicle_description": "GWB Class (GT46C-ACe II)"
        },
        {
            "loco_number": "NR84",
            "current_operator": "Pacific National",
            "vehicle_description": "NR Class"
        },
        {
            "loco_number": "DL43",
            "current_operator": "Aurizon",
            "vehicle_description": "DL Class"
        }
    ]

# =========================
# Insert / update logic
# =========================
def upsert_locos(scraped_locos):
    conn = get_db()
    cur = conn.cursor()

    blocked = load_blocked_locos()
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    added_count = 0
    updated_count = 0
    skipped_blocked = 0
    skipped_blank = 0

    for row in scraped_locos:
        loco_number = normalise_loco(row.get("loco_number"))
        current_operator = clean_text(row.get("current_operator"))
        vehicle_description = clean_text(row.get("vehicle_description"))

        if not loco_number:
            skipped_blank += 1
            continue

        if loco_number in blocked:
            skipped_blocked += 1
            continue

        cur.execute(
            "SELECT id, current_operator, vehicle_description, date_time_added FROM locos WHERE loco_number = ?",
            (loco_number,)
        )
        existing = cur.fetchone()

        if existing:
            # Update operator/description if changed, but keep original date_time_added
            if (
                clean_text(existing["current_operator"]) != current_operator
                or clean_text(existing["vehicle_description"]) != vehicle_description
            ):
                cur.execute("""
                    UPDATE locos
                    SET current_operator = ?, vehicle_description = ?
                    WHERE loco_number = ?
                """, (current_operator, vehicle_description, loco_number))
                updated_count += 1
        else:
            cur.execute("""
                INSERT INTO locos (
                    loco_number,
                    current_operator,
                    vehicle_description,
                    date_time_added
                ) VALUES (?, ?, ?, ?)
            """, (loco_number, current_operator, vehicle_description, now_str))
            added_count += 1

    conn.commit()
    conn.close()

    return {
        "added": added_count,
        "updated": updated_count,
        "skipped_blocked": skipped_blocked,
        "skipped_blank": skipped_blank
    }

# =========================
# Export spreadsheet
# =========================
def export_loco_spreadsheet():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            loco_number,
            current_operator,
            vehicle_description,
            date_time_added
        FROM locos
        ORDER BY loco_number COLLATE NOCASE ASC
    """)
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()

    # Remove default sheet and create proper ones
    default_ws = wb.active
    wb.remove(default_ws)

    ws_locos = wb.create_sheet("Locos")
    ws_blocked = wb.create_sheet("Blocked")
    ws_print = wb.create_sheet("Print View")
    ws_info = wb.create_sheet("Instructions")

    # -------------------------
    # Locos sheet
    # -------------------------
    loco_headers = [
        "Loco Number",
        "Current Operator",
        "Vehicle Description",
        "Date/Time Added",
    ]
    ws_locos.append(loco_headers)

    for cell in ws_locos[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws_locos.append([
            row["loco_number"],
            row["current_operator"],
            row["vehicle_description"],
            row["date_time_added"],
        ])

    # Set widths
    widths = {
        1: 18,
        2: 24,
        3: 38,
        4: 22,
    }
    for col_idx, width in widths.items():
        ws_locos.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws_locos.freeze_panes = "A2"

    # -------------------------
    # Blocked sheet
    # -------------------------
    ws_blocked.append(["Blocked Loco Number"])
    ws_blocked["A1"].font = Font(bold=True)
    ws_blocked.column_dimensions["A"].width = 22

    blocked = sorted(load_blocked_locos())
    for loco in blocked:
        ws_blocked.append([loco])

    # -------------------------
    # Print View sheet
    # -------------------------
    print_headers = [
        "Loco Number",
        "Current Operator",
        "Vehicle Description",
        "Date/Time Added",
    ]
    ws_print.append(print_headers)

    for cell in ws_print[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws_print.append([
            row["loco_number"],
            row["current_operator"],
            row["vehicle_description"],
            row["date_time_added"],
        ])

    for col_idx, width in widths.items():
        ws_print.column_dimensions[get_column_letter(col_idx)].width = width

    ws_print.freeze_panes = "A2"

    # Print settings
    ws_print.page_setup.orientation = "landscape"
    ws_print.page_setup.fitToWidth = 1
    ws_print.page_setup.fitToHeight = False
    ws_print.print_title_rows = "1:1"

    # -------------------------
    # Instructions sheet
    # -------------------------
    instructions = [
        "How to use this file:",
        "",
        "1. Locos sheet = your current live loco database.",
        "2. Blocked sheet = locos you never want re-added.",
        "3. Print View sheet = clean printable version.",
        "4. To permanently exclude a loco, add its number to blocked_locos.txt in the backend.",
        "5. This workbook is automatically rebuilt by the updater.",
    ]
    for line in instructions:
        ws_info.append([line])
    ws_info.column_dimensions["A"].width = 90

    # OpenPyXL saves xlsx workbooks directly to disk.
    wb.save(XLSX_PATH)

# =========================
# Main run
# =========================
def main():
    init_db()

    # 1. Scrape latest locos
    scraped_locos = scrape_latest_locos()

    # 2. Update DB
    stats = upsert_locos(scraped_locos)

    # 3. Export spreadsheet
    export_loco_spreadsheet()

    # 4. Console log
    print("Loco update complete")
    print(json.dumps(stats, indent=2))

if __name__ == "__main__":
    main()
