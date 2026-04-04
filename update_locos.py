import os
import json
import datetime
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

TRAINS_FILE = "trains.json"
LOCOS_FILE = "locos.json"
HISTORY_FILE = "loco_history.json"
EXPORT_FILE = "loco_export.csv"
SUMMARY_FILE = "loco_summary.txt"
BLOCKED_FILE = "blocked_locos.txt"
XLSX_DIR = os.path.join("static", "downloads")
XLSX_FILE = os.path.join(XLSX_DIR, "loco_database.xlsx")


def load_json(filename, default):
    if os.path.exists(filename):
        try:
            with open(filename, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"❌ Failed to load {filename}: {e}")
    return default


def save_json(filename, data):
    try:
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"❌ Failed to save {filename}: {e}")


def ensure_blocked_file():
    if not os.path.exists(BLOCKED_FILE):
        with open(BLOCKED_FILE, "w", encoding="utf-8") as f:
            f.write("# One loco number per line. These will never be re-added.\n")


def load_blocked_locos():
    ensure_blocked_file()
    blocked = set()

    try:
        with open(BLOCKED_FILE, "r", encoding="utf-8") as f:
            for raw_line in f:
                line = raw_line.strip()
                if not line or line.startswith("#"):
                    continue
                blocked.add(line.upper())
    except Exception as e:
        print(f"❌ Failed to read {BLOCKED_FILE}: {e}")

    return blocked


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalise_loco(value):
    return clean_text(value).upper()


def best_operator(train):
    return clean_text(
        train.get("current_operator")
        or train.get("operator")
        or train.get("operator_name")
        or ""
    )


def best_vehicle_description(train):
    # TrainFinder data in this repo usually does not include a proper dedicated
    # vehicle description field, so we keep the best available text.
    return clean_text(
        train.get("vehicle_description")
        or train.get("vehicleDescription")
        or train.get("description")
        or ""
    )


def export_to_csv(locos):
    fieldnames = [
        "Loco Number",
        "Current Operator",
        "Vehicle Description",
        "Date/Time Added",
        "Last Seen",
        "Last Date",
        "Last Time",
        "Latitude",
        "Longitude",
        "Speed",
        "Origin",
        "Destination",
        "Total Sightings",
        "cId",
        "servId",
        "trKey",
    ]

    active_locos = {k: v for k, v in locos.items() if not v.get("blocked", False)}

    try:
        with open(EXPORT_FILE, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            for loco_id in sorted(active_locos.keys()):
                data = active_locos[loco_id]
                if not isinstance(data, dict):
                    continue

                writer.writerow({
                    "Loco Number": loco_id,
                    "Current Operator": data.get("current_operator", ""),
                    "Vehicle Description": data.get("vehicle_description", ""),
                    "Date/Time Added": data.get("first_seen", "Unknown"),
                    "Last Seen": data.get("last_seen", ""),
                    "Last Date": data.get("last_date", ""),
                    "Last Time": data.get("last_time", ""),
                    "Latitude": data.get("last_location", {}).get("lat", ""),
                    "Longitude": data.get("last_location", {}).get("lon", ""),
                    "Speed": data.get("last_speed", 0),
                    "Origin": data.get("last_origin", ""),
                    "Destination": data.get("last_destination", ""),
                    "Total Sightings": data.get("total_sightings", 0),
                    "cId": data.get("cId", ""),
                    "servId": data.get("servId", ""),
                    "trKey": data.get("trKey", ""),
                })

        print(f"✅ Exported {len(active_locos)} active locos to {EXPORT_FILE}")
        print(f"📁 CSV path: {os.path.abspath(EXPORT_FILE)}")
        return True

    except Exception as e:
        print(f"❌ Failed to export CSV: {e}")
        return False


def create_summary(locos):
    active_locos = {k: v for k, v in locos.items() if not v.get("blocked", False)}
    blocked_count = sum(1 for _, v in locos.items() if isinstance(v, dict) and v.get("blocked", False))

    try:
        with open(SUMMARY_FILE, "w", encoding="utf-8") as f:
            f.write("=" * 80 + "\n")
            f.write(f"LOCO DATABASE SUMMARY - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Active Locomotives Tracked: {len(active_locos)}\n")
            f.write(f"Blocked Locomotives: {blocked_count}\n\n")

            sorted_locos = sorted(
                active_locos.items(),
                key=lambda x: x[1].get("last_seen", ""),
                reverse=True
            )

            for loco_id, data in sorted_locos[:50]:
                f.write(f"\n{'=' * 40}\n")
                f.write(f"LOCO: {loco_id}\n")
                f.write(f"{'=' * 40}\n")
                f.write(f"  Date/Time Added: {data.get('first_seen', 'Unknown')}\n")
                f.write(f"  Last Seen:       {data.get('last_seen', 'Unknown')}\n")
                f.write(f"  Operator:        {data.get('current_operator', '') or 'Unknown'}\n")
                f.write(f"  Vehicle Desc:    {data.get('vehicle_description', '') or 'Unknown'}\n")
                f.write(f"  Sightings:       {data.get('total_sightings', 0)}\n")
                f.write(
                    f"  Location:        "
                    f"({data.get('last_location', {}).get('lat', 'N/A')}, "
                    f"{data.get('last_location', {}).get('lon', 'N/A')})\n"
                )
                f.write(f"  Speed:           {data.get('last_speed', 0)} km/h\n")

                if data.get("last_origin") or data.get("last_destination"):
                    f.write(
                        f"  Route:           "
                        f"{data.get('last_origin', '?')} → {data.get('last_destination', '?')}\n"
                    )

        print(f"✅ Summary saved to {SUMMARY_FILE}")
        print(f"📁 Summary path: {os.path.abspath(SUMMARY_FILE)}")

    except Exception as e:
        print(f"❌ Failed to create summary: {e}")


def export_to_excel(locos):
    os.makedirs(XLSX_DIR, exist_ok=True)

    active_items = [
        (loco_id, data)
        for loco_id, data in sorted(locos.items(), key=lambda x: x[0].upper())
        if isinstance(data, dict) and not data.get("blocked", False)
    ]
    blocked_locos = sorted(load_blocked_locos())

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_locos = wb.create_sheet("Locos")
    ws_blocked = wb.create_sheet("Blocked")
    ws_print = wb.create_sheet("Print View")
    ws_info = wb.create_sheet("Instructions")

    headers = [
        "Loco Number",
        "Current Operator",
        "Vehicle Description",
        "Date/Time Added",
    ]

    # Locos
    ws_locos.append(headers)
    for cell in ws_locos[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for loco_id, data in active_items:
        ws_locos.append([
            loco_id,
            data.get("current_operator", ""),
            data.get("vehicle_description", ""),
            data.get("first_seen", ""),
        ])

    # Blocked
    ws_blocked.append(["Blocked Loco Number"])
    ws_blocked["A1"].font = Font(bold=True)
    for loco in blocked_locos:
        ws_blocked.append([loco])

    # Print view
    ws_print.append(headers)
    for cell in ws_print[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for loco_id, data in active_items:
        ws_print.append([
            loco_id,
            data.get("current_operator", ""),
            data.get("vehicle_description", ""),
            data.get("first_seen", ""),
        ])

    # Instructions
    instructions = [
        "This workbook is rebuilt automatically by update_locos.py.",
        "",
        "Locos = active loco database ready to view or filter.",
        "Blocked = loco numbers stored in blocked_locos.txt and excluded from re-entry.",
        "Print View = print-friendly copy of the active loco list.",
        "",
        "Important:",
        "- Date/Time Added comes from first_seen and is preserved for existing locos.",
        "- Current Operator and Vehicle Description only fill when the scraper provides those fields.",
        "- If TrainFinder leaves those blank, the workbook will leave them blank too.",
    ]
    for line in instructions:
        ws_info.append([line])

    widths = {
        1: 18,
        2: 24,
        3: 38,
        4: 22,
    }

    for ws in (ws_locos, ws_print):
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"

    ws_blocked.column_dimensions["A"].width = 24
    ws_info.column_dimensions["A"].width = 110

    ws_print.page_setup.orientation = "landscape"
    ws_print.page_setup.fitToWidth = 1
    ws_print.page_setup.fitToHeight = False
    ws_print.print_title_rows = "1:1"

    wb.save(XLSX_FILE)
    print(f"✅ Excel workbook saved to {XLSX_FILE}")


def update_loco_database():
    print("=" * 60)
    print("🚂 LOCO DATABASE UPDATER")
    print("=" * 60)

    if not os.path.exists(TRAINS_FILE):
        print(f"❌ {TRAINS_FILE} not found")
        return

    try:
        with open(TRAINS_FILE, "r", encoding="utf-8") as f:
            train_data = json.load(f)
    except Exception as e:
        print(f"❌ Failed to read {TRAINS_FILE}: {e}")
        return

    trains = train_data.get("trains", [])
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    date = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    print(f"\n📊 Processing {len(trains)} trains...")

    locos = load_json(LOCOS_FILE, {})
    history = load_json(HISTORY_FILE, {"locos": {}, "updates": []})
    blocked = load_blocked_locos()

    if not isinstance(locos, dict):
        locos = {}

    if not history or not isinstance(history, dict):
        history = {"locos": {}, "updates": []}

    history.setdefault("locos", {})
    history.setdefault("updates", [])

    # Keep anything listed in blocked_locos.txt flagged as blocked even if it already exists
    for blocked_loco in blocked:
        if blocked_loco in locos and isinstance(locos[blocked_loco], dict):
            locos[blocked_loco]["blocked"] = True

    new_locos = 0
    updated_locos = 0
    skipped_blocked = 0
    skipped_blank = 0

    for train in trains:
        loco_id = train.get("loco") or train.get("train_name") or train.get("train_number") or train.get("id")
        loco_id = normalise_loco(loco_id)

        if not loco_id:
            skipped_blank += 1
            continue

        if loco_id in blocked:
            skipped_blocked += 1
            if loco_id in locos and isinstance(locos[loco_id], dict):
                locos[loco_id]["blocked"] = True
            continue

        existing = locos.get(loco_id, {}) if isinstance(locos.get(loco_id), dict) else {}

        current_operator = best_operator(train) or existing.get("current_operator", "")
        vehicle_description = best_vehicle_description(train) or existing.get("vehicle_description", "")

        loco_data = {
            "first_seen": existing.get("first_seen", timestamp),
            "last_seen": timestamp,
            "last_date": date,
            "last_time": time_str,
            "date_time_added": existing.get("date_time_added", existing.get("first_seen", timestamp)),
            "current_operator": current_operator,
            "vehicle_description": vehicle_description,
            "last_location": {
                "lat": train.get("lat"),
                "lon": train.get("lon")
            },
            "last_speed": train.get("speed", 0),
            "last_origin": train.get("origin", ""),
            "last_destination": train.get("destination", ""),
            "last_description": train.get("description", ""),
            "last_train_number": train.get("train_number", ""),
            "cId": train.get("cId", ""),
            "servId": train.get("servId", ""),
            "trKey": train.get("trKey", ""),
            "total_sightings": existing.get("total_sightings", 0) + 1,
            "blocked": False,
        }

        if loco_id in locos and isinstance(locos[loco_id], dict):
            updated_locos += 1
        else:
            new_locos += 1

        locos[loco_id] = loco_data

        if loco_id not in history["locos"]:
            history["locos"][loco_id] = []

        history["locos"][loco_id].append({
            "timestamp": timestamp,
            "lat": train.get("lat"),
            "lon": train.get("lon"),
            "speed": train.get("speed", 0)
        })

        if len(history["locos"][loco_id]) > 100:
            history["locos"][loco_id] = history["locos"][loco_id][-100:]

    history["updates"].append({
        "timestamp": timestamp,
        "trains_seen": len(trains),
        "active_locos_seen": len([1 for _, data in locos.items() if isinstance(data, dict) and not data.get("blocked", False)]),
        "blocked_skipped": skipped_blocked,
    })

    if len(history["updates"]) > 1000:
        history["updates"] = history["updates"][-1000:]

    save_json(LOCOS_FILE, locos)
    save_json(HISTORY_FILE, history)

    print(f"\n📊 Statistics:")
    print(f"   New locos: {new_locos}")
    print(f"   Updated locos: {updated_locos}")
    print(f"   Skipped blocked: {skipped_blocked}")
    print(f"   Skipped blank: {skipped_blank}")
    print(f"   Total locos stored: {len(locos)}")

    export_to_csv(locos)
    create_summary(locos)
    export_to_excel(locos)


if __name__ == "__main__":
    update_loco_database()
